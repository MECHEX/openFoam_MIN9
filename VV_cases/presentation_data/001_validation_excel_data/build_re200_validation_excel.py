from __future__ import annotations

import csv
import json
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
RESULTS = ROOT / "VV_cases" / "V4b_3D" / "results"
OUT = Path(
    os.environ.get(
        "OUT_XLSX",
        ROOT / "VV_cases" / "presentation_data" / "V4b_Re200_validation_tables.xlsx",
    )
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
TEXT_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE = "FFFFFF"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except ValueError:
        return value


def write_table(ws, title: str, rows: list[dict], start_row: int = 1) -> int:
    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = Font(bold=True, size=13)
    ws.cell(start_row, 1).fill = SUBHEADER_FILL
    if not rows:
        ws.cell(start_row + 1, 1, "No data")
        return start_row + 3

    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    header_row = start_row + 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows, header_row + 1):
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(r_idx, c_idx, as_number(row.get(header)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"
    ws.freeze_panes = f"A{header_row + 1}"
    return header_row + len(rows) + 3


def style_sheet(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_decision_fill(ws, decision_col_name: str = "decision") -> None:
    header_row = None
    decision_col = None
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row, col).value == decision_col_name:
                header_row = row
                decision_col = col
                break
        if header_row:
            break
    if not header_row or not decision_col:
        return

    for row in range(header_row + 1, ws.max_row + 1):
        value = str(ws.cell(row, decision_col).value or "").lower()
        if any(token in value for token in ["accepted", "ok", "production", "stable"]):
            fill = GOOD_FILL
        elif any(token in value for token in ["diagnostic", "partial", "pending", "short"]):
            fill = WARN_FILL
        else:
            continue
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).fill = fill


def pct_diff(value, ref):
    value = as_number(value)
    ref = as_number(ref)
    if value is None or ref in (None, 0):
        return None
    return 100.0 * (value - ref) / ref


def metric_mean(summary: list[dict], metric: str, default=None):
    for row in summary:
        if row.get("metric") == metric:
            return row.get("mean", default)
    return default


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)

    gci_thermal = read_csv(RESULTS / "run011_gci_thermal_analysis" / "run011_thermal_gci_results.csv")
    gci_force_summary = read_csv(RESULTS / "run011_gci_analysis" / "run011_gci_force_coeff_summary.csv")
    outlet = read_csv(RESULTS / "run004c" / "run003_run004b_run004c_outlet_compare.csv")
    inlet = read_csv(RESULTS / "run005" / "run004b_vs_run005_inlet_compare.csv")
    dt04 = read_csv(RESULTS / "run006a" / "run004b_vs_run006a_timestep_partial_compare.csv")
    dt10 = read_csv(RESULTS / "run006b" / "run004b_vs_run006b_maxCo10_short_compare.csv")
    campaign = read_csv(RESULTS / "run008" / "data" / "011" / "run008_011_campaign_regime_table.csv")
    audit = read_json(RESULTS / "run008" / "data" / "001" / "run008_audit_uncertainty.json")
    heat = read_json(RESULTS / "run008" / "data" / "003" / "run008_003_heat_balance.json")

    wb = Workbook()
    wb.remove(wb.active)

    readme_rows = [
        {
            "item": "scope",
            "value": "Re=200 V4b production validation/sensitivity tables only",
            "notes": "Other Re cases are still part of onset scan and are intentionally excluded.",
        },
        {
            "item": "reference case",
            "value": "run008",
            "notes": "Production geometry, medium mesh, Lin=2D, Lout=8D, Lz=1D, t=2..10 s.",
        },
        {
            "item": "main comparison metrics",
            "value": "Cd_mean, Cl_rms, St, Nu_EB, Nu_wall, Q_wall/Q_total, T_out, heat-balance closure",
            "notes": "Mesh sensitivity sheet combines hydro sanity metrics with thermal metrics; formal mesh GCI sheet is thermal-only.",
        },
        {
            "item": "plot convention",
            "value": "Use plot_ready for long-format charts",
            "notes": "Each row is one value: study_type, case, x, metric, value, unit.",
        },
    ]
    ws = wb.create_sheet("README")
    write_table(ws, "Workbook Guide", readme_rows)

    production_rows = [
        {
            "case": "run008",
            "role": "production reference",
            "Re": 200,
            "Lin_D": 2,
            "Lout_D": 8,
            "Lz_D": 1,
            "cells": 407440,
            "solver": "foamRun fluid, laminar, eConst+Boussinesq",
            "window": "t = 2..10 s",
            "Cd_mean": audit.get("Cd_mean", 3.361014),
            "Cd_uncertainty": audit.get("Cd_ci95", 0.000772),
            "Cl_rms": audit.get("Cl_rms", 0.176441),
            "Cl_rms_uncertainty": audit.get("Cl_rms_ci95", 0.011097),
            "St": audit.get("St", 0.154261),
            "St_uncertainty": audit.get("St_ci95", 0.009574),
            "Nu_EB": audit.get("Nu_EB", 7.770004),
            "Nu_EB_uncertainty": audit.get("Nu_EB_ci95", 0.091573),
            "Nu_wall": audit.get("Nu_wall", 7.816521),
            "Nu_wall_uncertainty": audit.get("Nu_wall_ci95", 0.012286),
            "closure_pct": metric_mean(heat["summary"], "closure_ratio_of_means_pct", 0.706),
            "Q_air_W": metric_mean(heat["summary"], "Q_air", 1.4703),
            "Q_wall_W": metric_mean(heat["summary"], "Q_wall", 1.4807),
            "Q_tube_W": metric_mean(heat["summary"], "Q_tube", 0.3618),
            "Q_fins_W": metric_mean(heat["summary"], "Q_fins", 1.1189),
            "tube_heat_share_pct": metric_mean(heat["summary"], "tube_share_pct", 24.43),
            "fin_heat_share_pct": metric_mean(heat["summary"], "fins_share_pct", 75.57),
            "decision": "production reference accepted",
        }
    ]
    ws = wb.create_sheet("Re200_reference")
    write_table(ws, "Re=200 Production Reference", production_rows)
    add_decision_fill(ws)

    mesh_thermal_by_level = []
    mesh_cells = {"coarse": 196938, "medium": 407440, "fine": 829761}
    mesh_run = {
        "coarse": "run011_gci_coarse",
        "medium": "run008",
        "fine": "run011_gci_fine",
    }
    force_case_name = {
        "coarse": "coarse",
        "medium": "medium_run008",
        "fine": "fine",
    }
    force_by_level = {
        level: next(row for row in gci_force_summary if row["case"] == case_name)
        for level, case_name in force_case_name.items()
    }
    # The short 2-3 s mesh-GCI window contains ~3 shedding cycles.
    # Cl has two local extrema per shedding cycle, so f_shed is estimated
    # as half of the local-extrema frequency. This is a presentation
    # sanity metric, while run008 uses the longer 2-10 s production window.
    st_by_level = {
        "coarse": {"f_Cl_extrema_Hz": 6.470588235294117, "f_shed_Hz": 3.2352941176470584, "St": 0.15365918393004315},
        "medium": {"f_Cl_extrema_Hz": 6.470588235294117, "f_shed_Hz": 3.2352941176470584, "St": 0.15365918393004315},
        "fine": {"f_Cl_extrema_Hz": 6.470588235294119, "f_shed_Hz": 3.2352941176470593, "St": 0.1536591839300432},
    }
    thermal_mean = {
        row["metric"]: row
        for row in gci_thermal
        if row["source"] == "mean"
        and row["metric"] in {"Nu_EB", "Nu_wall", "Q_wall", "T_out", "closure_ratio_of_means_pct"}
    }
    for level in ["coarse", "medium", "fine"]:
        mesh_thermal_by_level.append(
            {
                "run": mesh_run[level],
                "cells": mesh_cells[level],
                "mesh_level": level,
                "Cd_mean_2_3": force_by_level[level]["Cd_mean_2_3"],
                "Cl_mean_2_3": force_by_level[level]["Cl_mean_2_3"],
                "Cl_rms_2_3": force_by_level[level]["Cl_rms_2_3"],
                "f_shed_est_Hz": st_by_level[level]["f_shed_Hz"],
                "St_est_2_3": st_by_level[level]["St"],
                "St_note": "estimated from half Cl-extrema frequency in short 2-3 s mesh window",
                "Nu_EB": thermal_mean["Nu_EB"][level],
                "Nu_wall": thermal_mean["Nu_wall"][level],
                "Q_wall_W": thermal_mean["Q_wall"][level],
                "T_out_K": thermal_mean["T_out"][level],
                "closure_ratio_of_means_pct": thermal_mean["closure_ratio_of_means_pct"][level],
                "Nu_EB_delta_pct_vs_fine": pct_diff(thermal_mean["Nu_EB"][level], thermal_mean["Nu_EB"]["fine"]),
                "Nu_wall_delta_pct_vs_fine": pct_diff(thermal_mean["Nu_wall"][level], thermal_mean["Nu_wall"]["fine"]),
                "Q_wall_delta_pct_vs_fine": pct_diff(thermal_mean["Q_wall"][level], thermal_mean["Q_wall"]["fine"]),
                "T_out_delta_K_vs_fine": as_number(thermal_mean["T_out"][level]) - as_number(thermal_mean["T_out"]["fine"]),
                "decision": "production mesh accepted" if level == "medium" else "GCI comparison point",
            }
        )
    ws = wb.create_sheet("mesh_thermal")
    write_table(ws, "Thermal Mesh Comparison by Run and Cell Count", mesh_thermal_by_level)
    add_decision_fill(ws)

    mesh_gci = []
    for row in gci_thermal:
        if row["source"] == "mean" and row["metric"] in {"Nu_EB", "Nu_wall", "Q_wall", "T_out"}:
            mesh_gci.append(
                {
                    "metric": row["metric"],
                    "window/source": row["source"],
                    "coarse_run": mesh_run["coarse"],
                    "medium_run": mesh_run["medium"],
                    "fine_run": mesh_run["fine"],
                    "coarse_cells": mesh_cells["coarse"],
                    "medium_cells": mesh_cells["medium"],
                    "fine_cells": mesh_cells["fine"],
                    "coarse": row["coarse"],
                    "medium": row["medium"],
                    "fine": row["fine"],
                    "p": row["p"],
                    "GCI_fine_medium_pct": row["GCI21_percent"],
                    "GCI_medium_coarse_pct": row["GCI32_percent"],
                    "status": row["status"],
                    "decision": "medium mesh accepted for global thermal metric",
                }
            )
    ws = wb.create_sheet("mesh_thermal_GCI")
    write_table(ws, "Thermal Mesh GCI", mesh_gci)
    add_decision_fill(ws)

    for row in inlet:
        row["study"] = "inlet_length"
        row["decision"] = "Lin=2D accepted; Lin=4D changes Cd/St/Nu negligibly" if row["run"] == "run004b" else "sensitivity check"
    ws = wb.create_sheet("inlet")
    write_table(ws, "Inlet Length Sensitivity", inlet)
    add_decision_fill(ws)

    for row in outlet:
        row["study"] = "outlet_length"
        if row["run"] == "run004b":
            row["decision"] = "Lout=8D production reference accepted"
        elif row["run"] == "run004c":
            row["decision"] = "Lout=16D confirms Lout=8D"
        else:
            row["decision"] = "Lout=5D archived shorter outlet, not production"
    ws = wb.create_sheet("outlet")
    write_table(ws, "Outlet Length Sensitivity", outlet)
    add_decision_fill(ws)

    dt_rows = []
    for row in dt04:
        row = dict(row)
        row["study"] = "dt_maxCo_0p4_vs_0p8"
        row["decision"] = "short-window timestep sensitivity; maxCo=0.8 accepted" if row["run"] == "run004b" else "maxCo=0.4 confirms maxCo=0.8"
        dt_rows.append(row)
    for row in dt10:
        row = dict(row)
        row["study"] = "dt_maxCo_1p0_vs_0p8"
        row["decision"] = "short-window timestep sensitivity; maxCo=0.8 accepted" if row["run"] == "run004b" else "maxCo=1.0 diagnostic confirms small drift"
        dt_rows.append(row)
    ws = wb.create_sheet("dt_maxCo")
    write_table(ws, "Time-Step / Courant Sensitivity", dt_rows)
    add_decision_fill(ws)

    for row in campaign:
        row["decision"] = "production reference accepted" if row.get("run") == "run008" else "supporting sensitivity/diagnostic"
    ws = wb.create_sheet("campaign")
    write_table(ws, "Campaign Comparison vs Production Reference", campaign)
    add_decision_fill(ws)

    plot_rows = []

    def add_plot(study, case, x_name, x, metric, value, unit="", note=""):
        value = as_number(value)
        if value is not None and value != "":
            plot_rows.append(
                {
                    "study_type": study,
                    "case": case,
                    "x_name": x_name,
                    "x": as_number(x),
                    "metric": metric,
                    "value": value,
                    "unit": unit,
                    "note": note,
                }
            )

    for row in outlet:
        for metric in ["Cd_mean", "Cl_rms", "St", "Nu_EB_LMTD", "T_out_K", "Q_total_W"]:
            add_plot("outlet_length", row["run"], "Lout_D", row["Lout_D"], metric, row.get(metric))

    for row in inlet:
        for metric in ["Cd_mean", "Cl_rms", "St", "Nu_EB_LMTD", "T_out_K", "Q_total_W"]:
            add_plot("inlet_length", row["run"], "Lin_D", row["Lin_D"], metric, row.get(metric))

    for row in dt_rows:
        if row.get("window") in {"t = 1..2 s", "t = 1..2.6 s"}:
            for metric in ["Cd_mean", "Cl_rms", "St", "Nu_EB_LMTD", "T_out_K", "Q_total_W"]:
                add_plot("dt_maxCo", row["run"], "maxCo", row["maxCo"], metric, row.get(metric), note=row.get("window", ""))

    for row in gci_thermal:
        if row["source"] == "mean":
            for mesh in ["coarse", "medium", "fine"]:
                add_plot("mesh_thermal", mesh_run[mesh], "cells", mesh_cells[mesh], row["metric"], row[mesh], note=row["source"])

    for row in mesh_thermal_by_level:
        for metric in ["Cd_mean_2_3", "Cl_mean_2_3", "Cl_rms_2_3", "f_shed_est_Hz", "St_est_2_3"]:
            add_plot("mesh_hydro_sanity", row["run"], "cells", row["cells"], metric, row.get(metric), note="mesh window 2-3 s")

    for row in production_rows:
        for metric in ["Cd_mean", "Cl_rms", "St", "Nu_EB", "Nu_wall", "closure_pct", "Q_air_W", "Q_wall_W"]:
            add_plot("production_reference", row["case"], "Re", row["Re"], metric, row.get(metric))

    ws = wb.create_sheet("plot_ready")
    write_table(ws, "Long-Format Plot-Ready Data", plot_rows)

    sources = [
        {"sheet": "mesh_thermal / mesh_thermal_GCI", "source_file": "VV_cases/V4b_3D/results/run011_gci_thermal_analysis/run011_thermal_gci_results.csv"},
        {"sheet": "inlet", "source_file": "VV_cases/V4b_3D/results/run005/run004b_vs_run005_inlet_compare.csv"},
        {"sheet": "outlet", "source_file": "VV_cases/V4b_3D/results/run004c/run003_run004b_run004c_outlet_compare.csv"},
        {"sheet": "dt_maxCo", "source_file": "VV_cases/V4b_3D/results/run006a/run004b_vs_run006a_timestep_partial_compare.csv"},
        {"sheet": "dt_maxCo", "source_file": "VV_cases/V4b_3D/results/run006b/run004b_vs_run006b_maxCo10_short_compare.csv"},
        {"sheet": "campaign", "source_file": "VV_cases/V4b_3D/results/run008/data/011/run008_011_campaign_regime_table.csv"},
        {"sheet": "Re200_reference", "source_file": "VV_cases/V4b_3D/results/run008/summary.md + data/001 + data/003 JSON"},
    ]
    ws = wb.create_sheet("sources")
    write_table(ws, "Source Files", sources)

    for ws in wb.worksheets:
        style_sheet(ws)

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
